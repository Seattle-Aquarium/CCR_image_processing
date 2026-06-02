#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
model_comparison_workbook_w_summary.py

Create a human-readable comparison workbook across multiple classification models using:
  - Ultralytics results.csv  (training / validation per-epoch metrics)
  - Per-label metrics report CSV (Class, Precision, Recall, F1 Score, …)

Sheets produced
───────────────
• Executive Summary  – overall winner, accuracy/macro-F1 per model, overfitting
                       verdict, per-category winner table, head-to-head win counts
• training_summary   – best epoch, top-1/top-5 accuracy, val loss, overfitting gap
• macro_summary      – macro P/R/F1/balanced-accuracy per model
                       (zero-sample classes excluded; weighted metrics omitted —
                        biased toward majority classes in imbalanced datasets)
• vs_<baseline>__<model>  – per-class Precision/Recall/F1 with Δ% vs baseline,
                             winner per class, low-sample flags
• <a>_vs_<b>_vs_<baseline>  – three-model side-by-side with bottleneck analysis
"""

from __future__ import annotations

import os
import sys
from dataclasses import dataclass
from typing import List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule


# ── colour palette ────────────────────────────────────────────────────────────
C_GREEN_BG  = "C6EFCE"
C_GREEN_FG  = "276221"
C_RED_BG    = "FFC7CE"
C_RED_FG    = "9C0006"
C_ORANGE_BG = "FFEB9C"
C_ORANGE_FG = "9C5700"
C_BLUE_HD   = "1F4E79"
C_BLUE_LT   = "D6E4F0"
C_GREY      = "F2F2F2"
C_WHITE     = "FFFFFF"

# F1 below this value is flagged as needing attention
LOW_F1_THRESHOLD   = 0.60
# Classes with fewer than this many test samples get a low-confidence flag
LOW_SAMPLE_MIN     = 10
# Two macro-F1 values within this tolerance are treated as tied
F1_TIE_TOLERANCE   = 0.001


# ── data classes ──────────────────────────────────────────────────────────────
@dataclass
class ModelFiles:
    name: str
    results_csv: str
    metrics_csv: str


# ── CSV helpers ───────────────────────────────────────────────────────────────
def _read_csv(path: str) -> pd.DataFrame:
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")
    df = pd.read_csv(path)
    df.columns = [c.strip() for c in df.columns]
    return df


def _coerce_numeric(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    for c in cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


# ── training summary ──────────────────────────────────────────────────────────
def summarize_results(results_df: pd.DataFrame) -> dict:
    if "epoch" not in results_df.columns:
        raise ValueError("results.csv missing required column: epoch")

    df = results_df.copy()
    df["epoch"] = pd.to_numeric(df["epoch"], errors="coerce")
    df = df.dropna(subset=["epoch"])
    df["epoch"] = df["epoch"].astype(int)

    best_by, best_row = None, None
    if "val/loss" in df.columns:
        best_by  = "val/loss (min)"
        best_row = df.loc[pd.to_numeric(df["val/loss"], errors="coerce").idxmin()]
    elif "metrics/accuracy_top1" in df.columns:
        best_by  = "accuracy_top1 (max)"
        best_row = df.loc[pd.to_numeric(df["metrics/accuracy_top1"], errors="coerce").idxmax()]
    else:
        best_by  = "epoch (last)"
        best_row = df.iloc[-1]

    def _get(col):
        return float(best_row[col]) if col in df.columns and pd.notna(best_row[col]) else None

    gap = None
    if "train/loss" in df.columns and "val/loss" in df.columns:
        tl, vl = _get("train/loss"), _get("val/loss")
        if tl is not None and vl is not None:
            gap = vl - tl

    overfit_verdict, overfit_suggestion = _overfit_analysis(gap, best_row, df)

    epochs_total = int(df["epoch"].max()) + 1 if df["epoch"].min() == 0 else int(df["epoch"].max())
    return {
        "epochs_total":          epochs_total,
        "best_by":               best_by,
        "best_epoch":            int(best_row["epoch"]),
        "best_val_loss":         _get("val/loss"),
        "best_top1_acc":         _get("metrics/accuracy_top1"),
        "best_top5_acc":         _get("metrics/accuracy_top5"),
        "gap_val_minus_train":   gap,
        "overfit_verdict":       overfit_verdict,
        "training_suggestion":   overfit_suggestion,
    }


def _overfit_analysis(gap, best_row, df) -> tuple[str, str]:
    if gap is None:
        return "Unknown", "Not enough loss data to assess overfitting."

    epochs_trained = int(df["epoch"].max())
    best_ep        = int(best_row["epoch"])
    near_end_note  = (
        f" NOTE: Best epoch ({best_ep}) is near the end of training "
        f"({epochs_trained} epochs) — consider training longer."
        if best_ep >= int(0.95 * epochs_trained) else ""
    )

    if gap > 0.1:
        return (
            "Significant overfitting",
            f"Val–train loss gap > 0.10. Consider stronger regularisation (dropout, "
            f"weight decay), data augmentation, or early stopping before epoch "
            f"{best_ep}.{near_end_note}",
        )
    if gap > 0.01:
        return (
            "Mild overfitting",
            f"Val–train gap 0.01–0.10. Model generalises reasonably; try light "
            f"augmentation or a small weight-decay increase.{near_end_note}",
        )
    return (
        "Good generalisation",
        f"Val–train gap < 0.01 — model generalises well. Could experiment with "
        f"training longer or a lower LR schedule.{near_end_note}",
    )


# ── metrics summary ───────────────────────────────────────────────────────────
def summarize_metrics(metrics_df: pd.DataFrame) -> dict:
    """
    Compute macro aggregates from per-class metrics.

    Zero-sample classes (e.g. 'background') are excluded from every macro
    average so they don't depress scores.  Weighted metrics are intentionally
    omitted — for imbalanced multi-class datasets they are dominated by the
    most common categories and mask poor performance on rare classes.
    """
    if "Class" not in metrics_df.columns:
        raise ValueError("metrics_report.csv missing required column: Class")

    df = metrics_df.copy()
    df["Class"] = df["Class"].astype(str).str.strip()
    df = _coerce_numeric(df, ["Total Samples", "Precision", "Recall",
                               "F1 Score", "Balanced Accuracy"])

    # Exclude zero-sample classes from aggregates
    support = df["Total Samples"] if "Total Samples" in df.columns else pd.Series([1] * len(df))
    valid = df[support > 0].copy()

    def macro(col):
        return float(valid[col].mean()) if col in valid.columns else None

    low_f1 = (
        valid.loc[valid["F1 Score"] < LOW_F1_THRESHOLD, "Class"].tolist()
        if "F1 Score" in valid.columns else []
    )
    low_sample = (
        df.loc[df["Total Samples"] < LOW_SAMPLE_MIN, "Class"].tolist()
        if "Total Samples" in df.columns else []
    )

    return {
        "n_classes":               int(df["Class"].nunique()),
        "n_classes_with_samples":  int(valid["Class"].nunique()),
        "macro_precision":         macro("Precision"),
        "macro_recall":            macro("Recall"),
        "macro_f1":                macro("F1 Score"),
        "macro_balanced_accuracy": macro("Balanced Accuracy"),
        "low_f1_classes":          low_f1,
        "low_sample_classes":      low_sample,
    }


# ── overall winner ─────────────────────────────────────────────────────────────
def _overall_winner(model_names: list, training_summaries: dict,
                    metrics_summaries: dict) -> tuple[str, str]:
    """
    Return (winner_name, reason).
    Primary criterion: macro F1.  Tiebreaker: top-1 accuracy.
    Mirrors the logic in ML_metrics_comparison.py.
    """
    scored = [
        (n,
         metrics_summaries[n].get("macro_f1") or 0.0,
         training_summaries[n].get("best_top1_acc") or 0.0)
        for n in model_names
    ]
    scored.sort(key=lambda x: (x[1], x[2]), reverse=True)

    winner, best_f1, _ = scored[0]
    if len(scored) > 1:
        second_f1 = scored[1][1]
        if abs(best_f1 - second_f1) < F1_TIE_TOLERANCE:
            return winner, "macro F1 tied; decided by top-1 accuracy"
        delta = best_f1 - second_f1
        return winner, f"highest macro F1 ({best_f1:.4f}, +{delta:.4f} over next best)"
    return winner, f"only model evaluated (macro F1 {best_f1:.4f})"


# ── per-label delta table ─────────────────────────────────────────────────────
def per_label_delta(
    baseline_metrics: pd.DataFrame,
    other_metrics: pd.DataFrame,
    baseline_name: str,
    other_name: str,
) -> pd.DataFrame:
    """
    Per-class comparison of Precision / Recall / F1 Score only.
    Balanced Accuracy is excluded (redundant with F1 for this use case).
    Only Δ% columns are produced — absolute deltas are omitted as redundant.
    Classes with fewer than LOW_SAMPLE_MIN test samples are flagged.
    Rows are sorted by absolute F1 Δ% (largest gaps first).
    """
    base = baseline_metrics.copy()
    oth  = other_metrics.copy()
    base["Class"] = base["Class"].astype(str).str.strip()
    oth["Class"]  = oth["Class"].astype(str).str.strip()

    # Only compare P / R / F1 — drop Balanced Accuracy per ML_metrics_comparison.py pattern
    compare_cols = [c for c in ["Precision", "Recall", "F1 Score"]
                    if c in base.columns and c in oth.columns]

    keep_base = ["Class"] + (["Total Samples"] if "Total Samples" in base.columns else [])
    base = base[keep_base + compare_cols].copy()
    oth  = oth[["Class"] + compare_cols].copy()

    merged = pd.merge(
        base, oth, on="Class",
        suffixes=(f"_{baseline_name}", f"_{other_name}"),
        how="outer",
    )
    merged = _coerce_numeric(merged, [c for c in merged.columns if c != "Class"])

    # Δ% only (no absolute delta)
    for col in compare_cols:
        delta = merged[f"{col}_{other_name}"] - merged[f"{col}_{baseline_name}"]
        merged[f"Δ%_{col}({other_name}-{baseline_name})"] = delta * 100

    # Winner per class by F1
    f1_base = merged.get(f"F1 Score_{baseline_name}")
    f1_oth  = merged.get(f"F1 Score_{other_name}")
    if f1_base is not None and f1_oth is not None:
        def _winner(row):
            a, b = row[f"F1 Score_{baseline_name}"], row[f"F1 Score_{other_name}"]
            if pd.isna(a) and pd.isna(b): return "neither"
            if pd.isna(a): return other_name
            if pd.isna(b): return baseline_name
            if abs(a - b) < F1_TIE_TOLERANCE: return "tie"
            return other_name if b > a else baseline_name
        merged["Winner_by_F1"] = merged.apply(_winner, axis=1)

    # Low-sample flag
    if "Total Samples" in merged.columns:
        merged["low_samples"] = merged["Total Samples"].apply(
            lambda n: "*" if pd.notna(n) and int(n) < LOW_SAMPLE_MIN else ""
        )

    # Sort by absolute F1 Δ% (biggest gaps first)
    delta_f1_col = f"Δ%_F1 Score({other_name}-{baseline_name})"
    if delta_f1_col in merged.columns:
        merged = merged.assign(_abs=merged[delta_f1_col].abs()) \
                       .sort_values("_abs", ascending=False) \
                       .drop(columns=["_abs"])

    return merged


# ── UIE side-by-side vs baseline ─────────────────────────────────────────────
def uie_side_by_side_vs_baseline(
    baseline_df, uie_a_df, uie_b_df,
    baseline_name, uie_a_name, uie_b_name,
) -> pd.DataFrame:
    """
    Three-model comparison: baseline vs two UIE variants.
    Shows F1/P/R for each model, Δ% vs baseline, bottleneck (P or R),
    UIE winner by F1, and direct UIE-vs-UIE Δ%.
    Balanced Accuracy and absolute delta columns are excluded.
    """
    def prep(df):
        d = df.copy()
        d.columns = [c.strip() for c in d.columns]
        d["Class"] = d["Class"].astype(str).str.strip()
        return _coerce_numeric(d, ["Total Samples", "F1 Score", "Precision", "Recall"])

    base = prep(baseline_df)
    a    = prep(uie_a_df)
    b    = prep(uie_b_df)

    for col in ["F1 Score", "Precision", "Recall"]:
        if col not in base.columns:
            raise ValueError(f"Baseline metrics CSV must contain '{col}' column.")

    keep_base = ["Class"] + (["Total Samples"] if "Total Samples" in base.columns else [])
    keep_base += ["F1 Score", "Precision", "Recall"]
    base = base[[c for c in keep_base if c in base.columns]].rename(columns={
        "F1 Score": f"F1 Score_{baseline_name}",
        "Precision": f"Precision_{baseline_name}",
        "Recall": f"Recall_{baseline_name}",
    })

    for df_uie, uname in [(a, uie_a_name), (b, uie_b_name)]:
        pass  # renamed below in merge

    keep_uie = [c for c in ["Class", "F1 Score", "Precision", "Recall"] if c in a.columns]
    a = a[keep_uie].rename(columns={
        "F1 Score": f"F1 Score_{uie_a_name}",
        "Precision": f"Precision_{uie_a_name}",
        "Recall": f"Recall_{uie_a_name}",
    })
    b = b[keep_uie].rename(columns={
        "F1 Score": f"F1 Score_{uie_b_name}",
        "Precision": f"Precision_{uie_b_name}",
        "Recall": f"Recall_{uie_b_name}",
    })

    merged = base.merge(a, on="Class", how="outer").merge(b, on="Class", how="outer")

    # Δ% only — no absolute deltas
    for metric in ["F1 Score", "Precision", "Recall"]:
        ca   = f"{metric}_{uie_a_name}"
        cb   = f"{metric}_{uie_b_name}"
        cbase = f"{metric}_{baseline_name}"
        if ca in merged.columns and cbase in merged.columns:
            merged[f"Δ%_{metric}({uie_a_name}-{baseline_name})"] = (
                (merged[ca] - merged[cbase]) * 100
            )
        if cb in merged.columns and cbase in merged.columns:
            merged[f"Δ%_{metric}({uie_b_name}-{baseline_name})"] = (
                (merged[cb] - merged[cbase]) * 100
            )
        if ca in merged.columns and cb in merged.columns:
            merged[f"Δ%_{metric}({uie_a_name}-{uie_b_name})"] = (
                (merged[ca] - merged[cb]) * 100
            )

    # Bottleneck: which metric degraded more relative to baseline
    def _bottleneck(row, model_name):
        rd  = row.get(f"Δ%_Recall({model_name}-{baseline_name})")
        prd = row.get(f"Δ%_Precision({model_name}-{baseline_name})")
        if pd.isna(rd) and pd.isna(prd): return None
        if pd.isna(rd):  return "Precision"
        if pd.isna(prd): return "Recall"
        if rd < prd: return "Recall"
        if prd < rd: return "Precision"
        return "Balanced"

    merged[f"bottleneck_({uie_a_name})"] = merged.apply(lambda r: _bottleneck(r, uie_a_name), axis=1)
    merged[f"bottleneck_({uie_b_name})"] = merged.apply(lambda r: _bottleneck(r, uie_b_name), axis=1)

    # UIE winner by F1
    def _winner(row):
        fa = row.get(f"F1 Score_{uie_a_name}")
        fb = row.get(f"F1 Score_{uie_b_name}")
        if pd.isna(fa) and pd.isna(fb): return "neither"
        if pd.isna(fa): return uie_b_name
        if pd.isna(fb): return uie_a_name
        if abs(fa - fb) < F1_TIE_TOLERANCE: return "tie"
        return uie_a_name if fa > fb else uie_b_name

    merged["UIE_winner_by_F1"] = merged.apply(_winner, axis=1)

    # Low-sample flag
    if "Total Samples" in merged.columns:
        merged["low_samples"] = merged["Total Samples"].apply(
            lambda n: "*" if pd.notna(n) and int(n) < LOW_SAMPLE_MIN else ""
        )

    # Column order
    cols = ["Class"]
    if "Total Samples"  in merged.columns: cols.append("Total Samples")
    if "low_samples"    in merged.columns: cols.append("low_samples")
    cols += [
        f"F1 Score_{baseline_name}", f"Precision_{baseline_name}", f"Recall_{baseline_name}",
        f"F1 Score_{uie_a_name}",
        f"Δ%_F1 Score({uie_a_name}-{baseline_name})",
        f"Δ%_Precision({uie_a_name}-{baseline_name})",
        f"Δ%_Recall({uie_a_name}-{baseline_name})",
        f"bottleneck_({uie_a_name})",
        f"F1 Score_{uie_b_name}",
        f"Δ%_F1 Score({uie_b_name}-{baseline_name})",
        f"Δ%_Precision({uie_b_name}-{baseline_name})",
        f"Δ%_Recall({uie_b_name}-{baseline_name})",
        f"bottleneck_({uie_b_name})",
        "UIE_winner_by_F1",
        f"Δ%_F1 Score({uie_a_name}-{uie_b_name})",
    ]
    cols   = [c for c in cols if c in merged.columns]
    merged = merged[cols]

    # Sort by absolute UIE-vs-UIE F1 delta (biggest gaps first)
    sort_col = f"Δ%_F1 Score({uie_a_name}-{uie_b_name})"
    if sort_col in merged.columns:
        merged = merged.assign(_abs=merged[sort_col].abs()) \
                       .sort_values("_abs", ascending=False) \
                       .drop(columns=["_abs"])

    return merged


# ── executive summary builder ─────────────────────────────────────────────────
def build_executive_summary(
    training_summaries: dict,
    metrics_summaries: dict,
    metrics_by_name: dict,
    model_names: list,
    baseline_name: str,
) -> list[list]:
    rows = []

    def h1(txt):   return [f"  {txt}"]
    def blank():   return [""]
    def row(*vals): return list(vals)

    rows.append(h1("MODEL COMPARISON  —  EXECUTIVE SUMMARY"))
    rows.append(blank())

    # ── overall winner ────────────────────────────────────────────────────────
    winner_name, winner_reason = _overall_winner(
        model_names, training_summaries, metrics_summaries
    )
    rows.append(h1("OVERALL WINNER"))
    rows.append(row("Best model", "Basis"))
    rows.append(row(winner_name, winner_reason))
    rows.append(row(
        "Note",
        "Primary criterion: macro F1 (equal weight per class; zero-sample classes "
        "excluded).  Tiebreaker: top-1 validation accuracy.",
    ))
    rows.append(blank())

    # ── section 1: overall metrics ────────────────────────────────────────────
    rows.append(h1("1 · OVERALL ACCURACY & MACRO METRICS"))
    rows.append(row(
        "Note",
        "Macro metrics give equal weight to every class — appropriate for "
        "imbalanced datasets.  Zero-sample classes (e.g. 'background') are "
        "excluded from all averages.  Weighted metrics are omitted as they are "
        "dominated by the most common classes.",
    ))
    rows.append(row("Model", "Top-1 Accuracy", "Macro Precision",
                    "Macro Recall", "Macro F1", "Macro Bal. Acc.",
                    "# Classes (with samples)", "# Low-F1 Classes (<0.60)"))
    for name in model_names:
        ts = training_summaries[name]
        ms = metrics_summaries[name]
        rows.append(row(
            name,
            f"{ts['best_top1_acc']:.3f}"         if ts['best_top1_acc']          is not None else "N/A",
            f"{ms['macro_precision']:.3f}"        if ms['macro_precision']        is not None else "N/A",
            f"{ms['macro_recall']:.3f}"           if ms['macro_recall']           is not None else "N/A",
            f"{ms['macro_f1']:.3f}"               if ms['macro_f1']               is not None else "N/A",
            f"{ms['macro_balanced_accuracy']:.3f}"if ms['macro_balanced_accuracy']is not None else "N/A",
            f"{ms['n_classes_with_samples']} / {ms['n_classes']}",
            len(ms['low_f1_classes']),
        ))
    rows.append(blank())

    # ── section 2: overfitting & training ────────────────────────────────────
    rows.append(h1("2 · TRAINING BEHAVIOUR & OVERFITTING"))
    rows.append(row("Model", "Total Epochs", "Best Epoch", "Val Loss @ Best",
                    "Val–Train Gap", "Verdict", "Suggestion"))
    for name in model_names:
        ts = training_summaries[name]
        rows.append(row(
            name,
            ts["epochs_total"],
            ts["best_epoch"],
            f"{ts['best_val_loss']:.4f}"        if ts['best_val_loss']        is not None else "N/A",
            f"{ts['gap_val_minus_train']:.5f}"  if ts['gap_val_minus_train']  is not None else "N/A",
            ts["overfit_verdict"],
            ts["training_suggestion"],
        ))
    rows.append(blank())

    # ── section 3: per-category winner ────────────────────────────────────────
    rows.append(h1("3 · PER-CATEGORY WINNER BY F1 SCORE"))
    rows.append(row(
        "Note",
        f"Sorted by F1 range (max − min across models) — most-contested classes "
        f"first.  * = fewer than {LOW_SAMPLE_MIN} test samples; treat with caution.",
    ))
    rows.append(row("Category", "Samples", *model_names, "Best Model", "Best F1",
                    "F1 Range", "Low-F1 in"))

    # build lookup
    all_classes = sorted(
        {cl for df in metrics_by_name.values()
         for cl in df["Class"].astype(str).str.strip().tolist()}
    )

    f1_lookup      = {}
    sample_lookup  = {}
    for name, df in metrics_by_name.items():
        d = df.copy()
        d["Class"] = d["Class"].astype(str).str.strip()
        d = _coerce_numeric(d, ["F1 Score", "Total Samples"])
        for _, r in d.iterrows():
            f1_lookup[(name, r["Class"])]     = r.get("F1 Score", float("nan"))
            sample_lookup[r["Class"]]          = r.get("Total Samples", 0)

    win_counts = {n: 0 for n in model_names}
    category_rows = []
    for cls in all_classes:
        f1s    = {n: f1_lookup.get((n, cls), float("nan")) for n in model_names}
        valid  = {n: v for n, v in f1s.items() if not pd.isna(v)}
        samples = int(sample_lookup.get(cls, 0))

        if valid:
            best_model = max(valid, key=valid.get)
            best_f1    = valid[best_model]
            f1_range   = max(valid.values()) - min(valid.values())
            win_counts[best_model] += 1
        else:
            best_model, best_f1, f1_range = "N/A", float("nan"), float("nan")

        low_in = [n for n, v in f1s.items() if not pd.isna(v) and v < LOW_F1_THRESHOLD]
        low_note = ", ".join(low_in) if low_in else ""
        flag     = "*" if samples < LOW_SAMPLE_MIN else ""

        category_rows.append({
            "cls": cls, "samples": samples, "f1s": f1s,
            "best_model": best_model, "best_f1": best_f1,
            "f1_range": f1_range, "low_note": low_note, "flag": flag,
        })

    # Sort by F1 range descending (most contested first)
    category_rows.sort(key=lambda x: x["f1_range"] if not pd.isna(x["f1_range"]) else -1,
                       reverse=True)

    for cr in category_rows:
        rows.append(row(
            f"{cr['cls']}{cr['flag']}",
            cr["samples"],
            *[f"{cr['f1s'][n]:.3f}" if not pd.isna(cr['f1s'][n]) else "—" for n in model_names],
            cr["best_model"],
            f"{cr['best_f1']:.3f}"   if not pd.isna(cr["best_f1"])   else "—",
            f"{cr['f1_range']:.3f}"  if not pd.isna(cr["f1_range"])  else "—",
            cr["low_note"],
        ))
    rows.append(blank())

    # ── section 4: head-to-head win counts ────────────────────────────────────
    rows.append(h1("4 · HEAD-TO-HEAD WIN COUNTS  (F1 per category)"))
    total_cats = len(all_classes)
    rows.append(row("Model", "Categories Won", "Win Rate"))
    for name in model_names:
        wc = win_counts[name]
        rows.append(row(name, wc, f"{wc / total_cats:.1%}" if total_cats else "N/A"))
    rows.append(blank())

    # ── section 5: categories needing attention ───────────────────────────────
    rows.append(h1(f"5 · CATEGORIES WITH F1 < {LOW_F1_THRESHOLD}  (need attention)"))
    rows.append(row(
        "Note",
        f"Classes flagged with * have fewer than {LOW_SAMPLE_MIN} test samples "
        f"— low F1 there may reflect data scarcity, not model weakness.",
    ))
    rows.append(row("Model", "Struggling Categories"))
    for name in model_names:
        ms  = metrics_summaries[name]
        low = ms["low_f1_classes"]
        rows.append(row(
            name,
            ", ".join(low) if low else "None — all categories above threshold",
        ))

    return rows


# ── openpyxl formatting helpers ────────────────────────────────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(hex_fg: str = C_WHITE, bold: bool = True, size: int = 11) -> Font:
    return Font(color=hex_fg, bold=bold, name="Arial", size=size)


def _thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _format_header_row(ws, row_idx: int, n_cols: int,
                       bg: str = C_BLUE_HD, fg: str = C_WHITE):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill      = _fill(bg)
        cell.font      = _font(fg)
        cell.border    = _thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autofit_columns(ws, min_width: int = 10, max_width: int = 50):
    for col_cells in ws.columns:
        length = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[col_cells[0].column_letter].width = max(min_width, min(length + 2, max_width))


def _color_scale_metric(ws, col_letter: str, first_row: int, last_row: int):
    """Green = high (good), red = low."""
    ws.conditional_formatting.add(
        f"{col_letter}{first_row}:{col_letter}{last_row}",
        ColorScaleRule(
            start_type="min",      start_color=C_RED_BG,
            mid_type="percentile", mid_value=50, mid_color="FFFFFF",
            end_type="max",        end_color=C_GREEN_BG,
        ),
    )


def _color_scale_delta(ws, col_letter: str, first_row: int, last_row: int):
    """Green = positive delta (improvement), red = negative delta."""
    ws.conditional_formatting.add(
        f"{col_letter}{first_row}:{col_letter}{last_row}",
        ColorScaleRule(
            start_type="min", start_color=C_RED_BG,
            mid_type="num",   mid_value=0, mid_color="FFFFFF",
            end_type="max",   end_color=C_GREEN_BG,
        ),
    )


def format_data_sheet(ws, metric_cols: list[str], delta_cols: list[str]):
    """Apply header formatting, alternating row tint, and color scales."""
    if ws.max_row < 2:
        return

    _format_header_row(ws, 1, ws.max_column)

    for r in range(2, ws.max_row + 1):
        row_fill = _fill(C_GREY) if r % 2 == 0 else _fill(C_WHITE)
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill   = row_fill
            cell.font   = Font(name="Arial", size=10)
            cell.border = _thin_border()

    headers = {
        ws.cell(row=1, column=c).value: get_column_letter(c)
        for c in range(1, ws.max_column + 1)
    }
    last_row = ws.max_row
    for col_name in metric_cols:
        if col_name in headers:
            _color_scale_metric(ws, headers[col_name], 2, last_row)
    for col_name in delta_cols:
        if col_name in headers:
            _color_scale_delta(ws, headers[col_name], 2, last_row)

    _autofit_columns(ws)
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False


# ── executive summary sheet writer ───────────────────────────────────────────
def write_executive_summary(wb, summary_rows: list[list]):
    ws = wb.create_sheet("Executive Summary", 0)
    ws.sheet_view.showGridLines = False

    section_fill = _fill(C_BLUE_HD)
    section_font = _font(C_WHITE, bold=True, size=12)
    subhdr_fill  = _fill(C_BLUE_LT)
    subhdr_font  = Font(color="1F4E79", bold=True, name="Arial", size=10)
    normal_font  = Font(name="Arial", size=10)
    wrap_align   = Alignment(wrap_text=True, vertical="top")

    col_widths: dict[int, int] = {}

    for r_idx, row_vals in enumerate(summary_rows, start=1):
        if not row_vals:
            continue
        first = str(row_vals[0]) if row_vals[0] else ""
        is_section = first.startswith("  ") and first.strip().isupper() or "·" in first
        is_blank   = first == ""

        for c_idx, val in enumerate(row_vals, start=1):
            cell           = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = wrap_align

            if is_section:
                cell.fill   = section_fill
                cell.font   = section_font
                cell.border = _thin_border()
            elif is_blank:
                ws.row_dimensions[r_idx].height = 6
            else:
                prev_first = str(summary_rows[r_idx - 2][0]) if r_idx >= 2 else ""
                is_subhdr  = (c_idx == 1 and
                              any(kw in str(val) for kw in ("Model", "Category", "Best model", "Note"))
                              and (prev_first.startswith("  ") or "·" in prev_first))
                cell.font   = subhdr_font if is_subhdr else normal_font
                cell.fill   = subhdr_fill if is_subhdr else _fill(C_WHITE)
                cell.border = _thin_border()

            cur_len = len(str(val)) if val else 0
            col_widths[c_idx] = max(col_widths.get(c_idx, 10), cur_len)

    for c_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max(width + 2, 12), 60)

    for r_idx, row_vals in enumerate(summary_rows, start=1):
        if row_vals and str(row_vals[0]).startswith("  "):
            ws.row_dimensions[r_idx].height = 22

    ws.freeze_panes = "A2"


# ── prompt helpers ─────────────────────────────────────────────────────────────
def prompt_models() -> List[ModelFiles]:
    models: List[ModelFiles] = []
    print("\nEnter models to compare (press Enter with no name when done).")
    print("  model name   — short label, e.g. 'hand', 'UNet', 'U-Shape'")
    print("  results.csv  — Ultralytics training output")
    print("  metrics_report.csv — per-class metrics (test split)\n")

    while True:
        name = input("Model name (Enter to finish): ").strip()
        if not name:
            break
        results_path = input(f"  {name} — path to results.csv: ").strip().strip('"')
        metrics_path = input(f"  {name} — path to metrics_report.csv: ").strip().strip('"')
        models.append(ModelFiles(name=name, results_csv=results_path, metrics_csv=metrics_path))

    if len(models) < 2:
        raise ValueError("Please enter at least two models to compare.")
    return models


# ── main ──────────────────────────────────────────────────────────────────────
def build_workbook(models: List[ModelFiles], baseline_name: str, out_dir: str) -> str:
    """Build model_comparison.xlsx from a list of ModelFiles; return the xlsx path.

    Importable, non-interactive entry point shared by the interactive CLI
    (`main`) and the automated pipeline (`compare_uie.py`).
    """
    names = [m.name for m in models]
    if len(models) < 2:
        raise ValueError("need at least two models to compare")
    if baseline_name not in names:
        raise ValueError(f"Baseline '{baseline_name}' not in: {names}")

    os.makedirs(out_dir, exist_ok=True)

    # ── load & summarise ──────────────────────────────────────────────────────
    training_summaries: dict = {}
    metrics_summaries:  dict = {}
    metrics_by_name:    dict = {}

    for m in models:
        res_df = _read_csv(m.results_csv)
        met_df = _read_csv(m.metrics_csv)
        met_df["Class"] = met_df["Class"].astype(str).str.strip()
        training_summaries[m.name] = summarize_results(res_df)
        metrics_summaries[m.name]  = summarize_metrics(met_df)
        metrics_by_name[m.name]    = met_df

    # training_summary sheet — exclude internal-only keys
    training_df = pd.DataFrame([
        {"model": n,
         "epochs_total":        training_summaries[n]["epochs_total"],
         "best_by":             training_summaries[n]["best_by"],
         "best_epoch":          training_summaries[n]["best_epoch"],
         "best_val_loss":       training_summaries[n]["best_val_loss"],
         "best_top1_acc":       training_summaries[n]["best_top1_acc"],
         "best_top5_acc":       training_summaries[n]["best_top5_acc"],
         "gap_val_minus_train": training_summaries[n]["gap_val_minus_train"],
         "overfit_verdict":     training_summaries[n]["overfit_verdict"],
         "training_suggestion": training_summaries[n]["training_suggestion"],
        }
        for n in names
    ]).sort_values("model")

    # macro_summary — only macro metrics (no weighted)
    macro_df = pd.DataFrame([
        {"model":                    n,
         "n_classes_with_samples":   metrics_summaries[n]["n_classes_with_samples"],
         "n_classes_total":          metrics_summaries[n]["n_classes"],
         "macro_precision":          metrics_summaries[n]["macro_precision"],
         "macro_recall":             metrics_summaries[n]["macro_recall"],
         "macro_f1":                 metrics_summaries[n]["macro_f1"],
         "macro_balanced_accuracy":  metrics_summaries[n]["macro_balanced_accuracy"],
        }
        for n in names
    ]).sort_values("model")

    # ── per-label delta sheets ────────────────────────────────────────────────
    base_df        = metrics_by_name[baseline_name]
    per_label_tbls = {}
    for m in models:
        if m.name == baseline_name:
            continue
        per_label_tbls[m.name] = per_label_delta(
            base_df, metrics_by_name[m.name], baseline_name, m.name
        )

    # ── UIE side-by-side sheet ────────────────────────────────────────────────
    non_baseline = [m.name for m in models if m.name != baseline_name]
    uie_vs_uie   = None
    uie_a_name = uie_b_name = None
    if len(non_baseline) >= 2:
        uie_a_name, uie_b_name = non_baseline[0], non_baseline[1]
        uie_vs_uie = uie_side_by_side_vs_baseline(
            baseline_df=metrics_by_name[baseline_name],
            uie_a_df=metrics_by_name[uie_a_name],
            uie_b_df=metrics_by_name[uie_b_name],
            baseline_name=baseline_name,
            uie_a_name=uie_a_name,
            uie_b_name=uie_b_name,
        )

    # ── write xlsx (first pass: data sheets) ─────────────────────────────────
    xlsx_path = os.path.join(out_dir, "model_comparison.xlsx")
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        training_df.to_excel(writer, sheet_name="training_summary", index=False)
        macro_df.to_excel(   writer, sheet_name="macro_summary",    index=False)

        for other_name, df in per_label_tbls.items():
            sheet = f"vs_{baseline_name}__{other_name}"[:31]
            df.to_excel(writer, sheet_name=sheet, index=False)

        if uie_vs_uie is not None:
            sheet_name = f"{uie_a_name}_vs_{uie_b_name}_vs_{baseline_name}"[:31]
            uie_vs_uie.to_excel(writer, sheet_name=sheet_name, index=False)

    # ── second pass: formatting ───────────────────────────────────────────────
    wb = load_workbook(xlsx_path)

    # training_summary
    ws_tr = wb["training_summary"]
    format_data_sheet(ws_tr,
        metric_cols=["best_top1_acc", "best_top5_acc"],
        delta_cols=["gap_val_minus_train"],
    )
    for col_cells in ws_tr.columns:
        hdr = ws_tr.cell(row=1, column=col_cells[0].column).value
        if hdr and "suggestion" in str(hdr).lower():
            ws_tr.column_dimensions[col_cells[0].column_letter].width = 60
            for cell in col_cells[1:]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    for r in range(2, ws_tr.max_row + 1):
        ws_tr.row_dimensions[r].height = 55

    # macro_summary — only macro metric columns (no weighted)
    ws_mac = wb["macro_summary"]
    format_data_sheet(ws_mac,
        metric_cols=["macro_precision", "macro_recall",
                     "macro_f1", "macro_balanced_accuracy"],
        delta_cols=[],
    )

    # per-label delta sheets
    for other_name in per_label_tbls:
        sheet = f"vs_{baseline_name}__{other_name}"[:31]
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        f1_cols = [
            ws.cell(row=1, column=c).value
            for c in range(1, ws.max_column + 1)
            if ws.cell(row=1, column=c).value
            and "F1" in str(ws.cell(row=1, column=c).value)
            and "Δ" not in str(ws.cell(row=1, column=c).value)
        ]
        delta_cols = [
            ws.cell(row=1, column=c).value
            for c in range(1, ws.max_column + 1)
            if ws.cell(row=1, column=c).value
            and "Δ%" in str(ws.cell(row=1, column=c).value)
        ]
        format_data_sheet(ws, metric_cols=f1_cols, delta_cols=delta_cols)

    # UIE side-by-side sheet
    if uie_vs_uie is not None:
        sheet_name = f"{uie_a_name}_vs_{uie_b_name}_vs_{baseline_name}"[:31]
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            f1_cols = [
                ws.cell(row=1, column=c).value
                for c in range(1, ws.max_column + 1)
                if ws.cell(row=1, column=c).value
                and "F1" in str(ws.cell(row=1, column=c).value)
                and "Δ" not in str(ws.cell(row=1, column=c).value)
            ]
            delta_cols = [
                ws.cell(row=1, column=c).value
                for c in range(1, ws.max_column + 1)
                if ws.cell(row=1, column=c).value
                and "Δ%" in str(ws.cell(row=1, column=c).value)
            ]
            format_data_sheet(ws, metric_cols=f1_cols, delta_cols=delta_cols)

    # Executive summary (inserted as first sheet)
    summary_rows = build_executive_summary(
        training_summaries, metrics_summaries, metrics_by_name, names, baseline_name
    )
    write_executive_summary(wb, summary_rows)

    wb.save(xlsx_path)
    print(f"\nSaved workbook:\n   {xlsx_path}")
    print("\nSheets:")
    for s in wb.sheetnames:
        print(f"   • {s}")
    return xlsx_path


def main():
    models = prompt_models()
    names = [m.name for m in models]

    baseline_name = input("\nBaseline model name (must match one you entered): ").strip()
    if baseline_name not in names:
        raise ValueError(f"Baseline '{baseline_name}' not in: {names}")

    out_dir = input("\nOutput folder (blank = current directory): ").strip().strip('"') or os.getcwd()
    build_workbook(models, baseline_name, out_dir)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\nERROR: {e}")
        sys.exit(1)
